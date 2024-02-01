# SPDX-License-Identifier: LGPL-2.1-or-later
# https://github.com/mdealencar/interarray

from pathlib import Path

import networkx as nx
import numpy as np
import scipy.io
import utm
import yaml
from openpyxl import load_workbook

from .geometric import make_graph_metrics, rotate
from .utils import NodeTagger


F = NodeTagger()


def utm_from_latlonstr(entries):
    utm_out = []
    for entry in entries.splitlines():
        label, lat, lon = entry.split(' ')
        coords = []
        for ll in (lat, lon):
            val, hemisphere = ll.split("'")
            deg, sec = val.split('°')
            coords.append((float(deg) + float(sec)/60) * (1 if hemisphere in ('N', 'E') else -1))
        utm_out.append((label, *utm.from_latlon(*coords)))
    return utm_out


def file2graph(filename, rotation=None, handle='file'):
    '''filename is a Matlab .mat file or an Excel
    spreadsheet in the proper format'''
    fpath = Path(filename)
    data = {}
    # read Excel xls file
    if fpath.suffix == '.xls':
        print('xls reading not implemented, save as .xlsx instead')
        return None
    # read wind power plant YAML file
    elif fpath.suffix == '.yaml':
        scrapped = yaml.safe_load(open(fpath, 'r', encoding='utf8'))
        for key, scrapped_key in (('WT coordinates', 'TURBINES'),
                                  ('OSS coordinates', 'SUBSTATIONS'),
                                  ('WF area limits', 'EXTENTS')):
            source = utm_from_latlonstr(scrapped[scrapped_key])
            labels, easting, northing, zone_number, zone_letter = zip(*source)
            xy = np.array((easting, northing))
            if rotation is not None:
                coords = rotate(xy.T, rotation).T
            else:
                coords = xy
            data[key] = coords, labels
    # read Excel xlsx file
    elif fpath.suffix == '.xlsx':
        wb = load_workbook(filename=fpath, read_only=True, data_only=True)
        for ws in wb.worksheets[:4]:
            key = ws['A1'].value
            if key not in ['WF area limits', 'OSS coordinates',
                           'WT coordinates', 'Forbidden Zones']:
                continue
            if key == 'Forbidden Zones':
                if ws['A3'].value is not None:
                    print('Forbidden Zones not yet implemented.')
                continue
            for cell, header in (('A2', 'x (m)'),
                                 ('B2', 'y (m)')):
                assert ws[cell].value == header
            xy = [(float(x.value), float(y.value)) for x, y in
                  ws.iter_rows(min_row=3, min_col=1, max_col=2) if
                  ((x.data_type == 'n') and
                  (y.data_type == 'n') and
                  (x.value is not None) and
                  (y.value is not None))]
            labels = [lab.value for lab, in
                      ws.iter_rows(min_row=3, min_col=3, max_col=3)
                      if lab.value is not None]
            if len(xy) != len(labels):
                labels = None
            if xy:
                if rotation is not None:
                    coords = rotate(np.array(xy), rotation).T
                else:
                    coords = np.array(tuple(zip(*xy)), dtype=float)
                data[key] = coords, labels
    # read Matlab mat file
    elif fpath.suffix == '.mat':
        windfarm = scipy.io.loadmat(fpath,
                                    struct_as_record=False,
                                    squeeze_me=True)['WindFarm']
        data['WT coordinates'] = (np.r_[[windfarm.Coord.x[1:]],
                                        [windfarm.Coord.y[1:]]],
                                  None)
        data['WF area limits'] = (np.r_[[windfarm.Area.xv],
                                        [windfarm.Area.yv]],
                                  None)
        # TODO: is the matlab data structure compatible with multiple OSS?
        data['OSS coordinates'] = (np.r_[[windfarm.Coord.x[0]],
                                         [windfarm.Coord.y[0]]],
                                   None)

    # TODO: add node labels/ids to graph properties, if given

    # build data structures
    WTcoords, WTlabels = data['WT coordinates']
    OSScoords, OSSlabels = data['OSS coordinates']
    boundary = data['WF area limits'][0].T
    N = WTcoords.shape[1]
    M = OSScoords.shape[1]
    # create networkx graph
    G = nx.Graph(M=M,
                 VertexC=np.vstack((WTcoords.T, OSScoords.T[::-1])),
                 boundary=boundary,
                 name=fpath.stem,
                 handle=handle)
    G.add_nodes_from(((n, {'label': F[n], 'type': 'wtg'})
                      for n in range(N)))
    G.add_nodes_from(((r, {'label': F[r], 'type': 'oss'})
                      for r in range(-M, 0)))
    make_graph_metrics(G)
    return G
